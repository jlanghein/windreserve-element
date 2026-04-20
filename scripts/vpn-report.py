#!/usr/bin/env python3
"""
WireGuard VPN Report Generator

Creates a professional Excel spreadsheet report of all VPN clients and users
from the wg-portal management system.

Usage:
    uvx --with openpyxl python scripts/vpn-report.py [--output FILE]

Requirements:
    - SSH access to caddy.helium (10.24.0.29) via sshpass
    - wg-portal credentials from secrets.env
"""

import argparse
import json
import subprocess
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Border, Side, Alignment, NamedStyle
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    print("Error: openpyxl not installed. Run with:")
    print("  uvx --with openpyxl python scripts/vpn-report.py")
    sys.exit(1)


# Colors - WindReserve corporate style
COLORS = {
    'primary': '2E7D32',      # Dark green
    'secondary': '66BB6A',    # Light green
    'accent': '1565C0',       # Blue
    'warning': 'FFA726',      # Orange
    'danger': 'EF5350',       # Red
    'light': 'E8F5E9',        # Very light green
    'dark': '1B5E20',         # Very dark green
    'white': 'FFFFFF',
    'gray_light': 'F5F5F5',
    'gray': 'BDBDBD',
}


def load_secrets():
    """Load secrets from secrets.env file."""
    secrets = {}
    secrets_file = Path(__file__).parent.parent / 'secrets.env'
    
    if not secrets_file.exists():
        print(f"Error: {secrets_file} not found")
        sys.exit(1)
    
    with open(secrets_file) as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                key, value = line.split('=', 1)
                secrets[key] = value
    
    return secrets


def fetch_vpn_data(secrets):
    """Fetch VPN data from wg-portal via SSH."""
    ssh_host = secrets.get('CADDY_HELIUM_HOST', '10.24.0.29')
    ssh_user = secrets.get('CADDY_HELIUM_USER', 'wradmin')
    ssh_pass = secrets.get('CADDY_HELIUM_PASS', '')
    wg_user = secrets.get('WGPORTAL_USER', '')
    wg_pass = secrets.get('WGPORTAL_PASS', '')
    wg_host = secrets.get('WGPORTAL_HOST', '10.21.254.2')
    wg_port = secrets.get('WGPORTAL_PORT', '8888')
    
    api_url = f"http://{wg_host}:{wg_port}"
    
    # Login to get session
    login_json = json.dumps({"username": wg_user, "password": wg_pass})
    # Escape for shell
    login_json_escaped = login_json.replace("'", "'\\''")
    
    login_cmd = [
        'sshpass', '-p', ssh_pass,
        'ssh', '-o', 'StrictHostKeyChecking=no', f'{ssh_user}@{ssh_host}',
        f"curl -s -i -X POST {api_url}/api/v0/auth/login -H 'Content-Type: application/json' -d '{login_json_escaped}'"
    ]
    
    result = subprocess.run(login_cmd, capture_output=True, text=True)
    
    # Extract session cookie
    session = None
    for line in result.stdout.split('\n'):
        if 'wgPortalSession=' in line:
            session = line.split('wgPortalSession=')[1].split(';')[0]
            break
    
    if not session:
        print("Error: Failed to login to wg-portal")
        print(result.stdout)
        print(result.stderr)
        sys.exit(1)
    
    # Fetch peers from both interfaces
    peers = []
    
    for iface, iface_name in [('d2cw', 'wg0 (MA VPN)'), ('d2cx', 'wg1 (IT VPN)')]:
        cmd = [
            'sshpass', '-p', ssh_pass,
            'ssh', '-o', 'StrictHostKeyChecking=no', f'{ssh_user}@{ssh_host}',
            f"curl -s -H 'Cookie: wgPortalSession={session}' '{api_url}/api/v0/peer/iface/{iface}/all'"
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True)
        
        try:
            iface_peers = json.loads(result.stdout)
            for peer in iface_peers:
                peer['Interface'] = iface_name
                peer['InterfaceShort'] = iface_name.split()[0]
            peers.extend(iface_peers)
        except json.JSONDecodeError:
            print(f"Warning: Failed to parse peers for {iface_name}")
    
    # Fetch users
    cmd = [
        'sshpass', '-p', ssh_pass,
        'ssh', '-o', 'StrictHostKeyChecking=no', f'{ssh_user}@{ssh_host}',
        f"curl -s -H 'Cookie: wgPortalSession={session}' '{api_url}/api/v0/user/all'"
    ]
    
    result = subprocess.run(cmd, capture_output=True, text=True)
    
    try:
        users = json.loads(result.stdout)
    except json.JSONDecodeError:
        print("Warning: Failed to parse users")
        users = []
    
    return peers, users


def create_styles():
    """Create Excel styles for the report."""
    styles = {}
    
    # Header style
    styles['header'] = NamedStyle(name='header')
    styles['header'].font = Font(bold=True, color=COLORS['white'], size=11)
    styles['header'].fill = PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid')
    styles['header'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    styles['header'].border = Border(
        left=Side(style='thin', color=COLORS['dark']),
        right=Side(style='thin', color=COLORS['dark']),
        top=Side(style='thin', color=COLORS['dark']),
        bottom=Side(style='thin', color=COLORS['dark'])
    )
    
    # Title style
    styles['title'] = NamedStyle(name='title')
    styles['title'].font = Font(bold=True, color=COLORS['dark'], size=24)
    styles['title'].alignment = Alignment(horizontal='left', vertical='center')
    
    # Subtitle style
    styles['subtitle'] = NamedStyle(name='subtitle')
    styles['subtitle'].font = Font(bold=False, color=COLORS['gray'], size=12, italic=True)
    styles['subtitle'].alignment = Alignment(horizontal='left', vertical='center')
    
    # Data cell style
    styles['data'] = NamedStyle(name='data')
    styles['data'].font = Font(size=10)
    styles['data'].alignment = Alignment(horizontal='left', vertical='center')
    styles['data'].border = Border(
        left=Side(style='thin', color=COLORS['gray']),
        right=Side(style='thin', color=COLORS['gray']),
        top=Side(style='thin', color=COLORS['gray']),
        bottom=Side(style='thin', color=COLORS['gray'])
    )
    
    # Enabled status
    styles['enabled'] = NamedStyle(name='enabled')
    styles['enabled'].font = Font(size=10, bold=True, color=COLORS['primary'])
    styles['enabled'].fill = PatternFill(start_color=COLORS['light'], end_color=COLORS['light'], fill_type='solid')
    styles['enabled'].alignment = Alignment(horizontal='center', vertical='center')
    styles['enabled'].border = Border(
        left=Side(style='thin', color=COLORS['gray']),
        right=Side(style='thin', color=COLORS['gray']),
        top=Side(style='thin', color=COLORS['gray']),
        bottom=Side(style='thin', color=COLORS['gray'])
    )
    
    # Disabled status
    styles['disabled'] = NamedStyle(name='disabled')
    styles['disabled'].font = Font(size=10, bold=True, color=COLORS['danger'])
    styles['disabled'].fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
    styles['disabled'].alignment = Alignment(horizontal='center', vertical='center')
    styles['disabled'].border = Border(
        left=Side(style='thin', color=COLORS['gray']),
        right=Side(style='thin', color=COLORS['gray']),
        top=Side(style='thin', color=COLORS['gray']),
        bottom=Side(style='thin', color=COLORS['gray'])
    )
    
    # MA VPN style
    styles['ma_vpn'] = NamedStyle(name='ma_vpn')
    styles['ma_vpn'].font = Font(size=10, bold=True, color=COLORS['primary'])
    styles['ma_vpn'].fill = PatternFill(start_color=COLORS['light'], end_color=COLORS['light'], fill_type='solid')
    styles['ma_vpn'].alignment = Alignment(horizontal='center', vertical='center')
    styles['ma_vpn'].border = Border(
        left=Side(style='thin', color=COLORS['gray']),
        right=Side(style='thin', color=COLORS['gray']),
        top=Side(style='thin', color=COLORS['gray']),
        bottom=Side(style='thin', color=COLORS['gray'])
    )
    
    # IT VPN style
    styles['it_vpn'] = NamedStyle(name='it_vpn')
    styles['it_vpn'].font = Font(size=10, bold=True, color=COLORS['accent'])
    styles['it_vpn'].fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
    styles['it_vpn'].alignment = Alignment(horizontal='center', vertical='center')
    styles['it_vpn'].border = Border(
        left=Side(style='thin', color=COLORS['gray']),
        right=Side(style='thin', color=COLORS['gray']),
        top=Side(style='thin', color=COLORS['gray']),
        bottom=Side(style='thin', color=COLORS['gray'])
    )
    
    # Stat box style
    styles['stat_value'] = NamedStyle(name='stat_value')
    styles['stat_value'].font = Font(bold=True, color=COLORS['dark'], size=36)
    styles['stat_value'].alignment = Alignment(horizontal='center', vertical='center')
    
    styles['stat_label'] = NamedStyle(name='stat_label')
    styles['stat_label'].font = Font(color=COLORS['gray'], size=11)
    styles['stat_label'].alignment = Alignment(horizontal='center', vertical='center')
    
    return styles


def create_report(peers, users, output_file):
    """Create the Excel report."""
    wb = Workbook()
    styles = create_styles()
    
    # Register styles
    for style in styles.values():
        try:
            wb.add_named_style(style)
        except ValueError:
            pass  # Style already exists
    
    # ========== DASHBOARD SHEET ==========
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    
    # Set column widths
    ws_dash.column_dimensions['A'].width = 3
    ws_dash.column_dimensions['B'].width = 25
    ws_dash.column_dimensions['C'].width = 15
    ws_dash.column_dimensions['D'].width = 15
    ws_dash.column_dimensions['E'].width = 15
    ws_dash.column_dimensions['F'].width = 15
    ws_dash.column_dimensions['G'].width = 25
    ws_dash.column_dimensions['H'].width = 15
    
    # Title
    ws_dash.merge_cells('B2:H2')
    ws_dash['B2'] = "WindReserve VPN Report"
    ws_dash['B2'].style = 'title'
    
    # Subtitle
    ws_dash.merge_cells('B3:H3')
    ws_dash['B3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_dash['B3'].style = 'subtitle'
    
    # Calculate statistics
    total_peers = len(peers)
    active_peers = sum(1 for p in peers if not p.get('Disabled', False))
    disabled_peers = total_peers - active_peers
    ma_vpn_peers = sum(1 for p in peers if p.get('InterfaceShort') == 'wg0')
    it_vpn_peers = sum(1 for p in peers if p.get('InterfaceShort') == 'wg1')
    total_users = len(users)
    active_users = sum(1 for u in users if not u.get('Disabled', False))
    
    # Unique users with peers
    unique_users_with_peers = len(set(p.get('UserIdentifier', '') for p in peers if p.get('UserIdentifier')))
    
    # Stats boxes - Row 5-7
    stats = [
        ('Total Peers', total_peers, COLORS['primary']),
        ('Active', active_peers, COLORS['secondary']),
        ('Disabled', disabled_peers, COLORS['danger']),
        ('MA VPN', ma_vpn_peers, COLORS['primary']),
        ('IT VPN', it_vpn_peers, COLORS['accent']),
    ]
    
    col = 2  # Start at column B
    for label, value, color in stats:
        cell_value = ws_dash.cell(row=6, column=col)
        cell_value.value = value
        cell_value.font = Font(bold=True, color=color, size=32)
        cell_value.alignment = Alignment(horizontal='center', vertical='center')
        
        cell_label = ws_dash.cell(row=7, column=col)
        cell_label.value = label
        cell_label.style = 'stat_label'
        
        # Add background box
        for r in [5, 6, 7, 8]:
            cell = ws_dash.cell(row=r, column=col)
            cell.fill = PatternFill(start_color=COLORS['gray_light'], end_color=COLORS['gray_light'], fill_type='solid')
            cell.border = Border(
                left=Side(style='thin', color=COLORS['gray']),
                right=Side(style='thin', color=COLORS['gray']),
                top=Side(style='thin', color=COLORS['gray']) if r == 5 else None,
                bottom=Side(style='thin', color=COLORS['gray']) if r == 8 else None
            )
        
        col += 1
    
    # User stats - Row 10-12
    ws_dash['B10'] = "User Statistics"
    ws_dash['B10'].font = Font(bold=True, size=14, color=COLORS['dark'])
    
    user_stats = [
        ('Total Users', total_users),
        ('Active Users', active_users),
        ('Users with Peers', unique_users_with_peers),
    ]
    
    col = 2
    for label, value in user_stats:
        cell_value = ws_dash.cell(row=12, column=col)
        cell_value.value = value
        cell_value.font = Font(bold=True, color=COLORS['accent'], size=24)
        cell_value.alignment = Alignment(horizontal='center', vertical='center')
        
        cell_label = ws_dash.cell(row=13, column=col)
        cell_label.value = label
        cell_label.style = 'stat_label'
        col += 1
    
    # Peers per user analysis
    ws_dash['B15'] = "Peers Per User (Top 10)"
    ws_dash['B15'].font = Font(bold=True, size=14, color=COLORS['dark'])
    
    user_peer_count = {}
    for peer in peers:
        user = peer.get('UserIdentifier', 'Unknown')
        if user:
            user_peer_count[user] = user_peer_count.get(user, 0) + 1
    
    sorted_users = sorted(user_peer_count.items(), key=lambda x: -x[1])[:10]
    
    ws_dash['B17'] = "User"
    ws_dash['C17'] = "Peer Count"
    ws_dash['B17'].style = 'header'
    ws_dash['C17'].style = 'header'
    
    for i, (user, count) in enumerate(sorted_users, start=18):
        ws_dash.cell(row=i, column=2).value = user.replace('@windreserve.de', '')
        ws_dash.cell(row=i, column=2).style = 'data'
        ws_dash.cell(row=i, column=3).value = count
        ws_dash.cell(row=i, column=3).style = 'data'
        ws_dash.cell(row=i, column=3).alignment = Alignment(horizontal='center')
    
    # ========== VPN CLIENTS SHEET ==========
    ws_clients = wb.create_sheet("VPN Clients")
    
    # Headers
    headers = ['#', 'Display Name', 'User', 'IP Address', 'VPN Interface', 'Status', 'Notes']
    for col, header in enumerate(headers, start=1):
        cell = ws_clients.cell(row=1, column=col)
        cell.value = header
        cell.style = 'header'
    
    # Set column widths
    ws_clients.column_dimensions['A'].width = 5
    ws_clients.column_dimensions['B'].width = 35
    ws_clients.column_dimensions['C'].width = 30
    ws_clients.column_dimensions['D'].width = 18
    ws_clients.column_dimensions['E'].width = 18
    ws_clients.column_dimensions['F'].width = 12
    ws_clients.column_dimensions['G'].width = 30
    
    # Sort peers: active first, then by interface, then by IP
    def sort_key(p):
        disabled = 1 if p.get('Disabled') else 0
        iface = 0 if p.get('InterfaceShort') == 'wg0' else 1
        addr = p.get('Addresses', [''])[0] if isinstance(p.get('Addresses'), list) else p.get('Addresses', '')
        ip_num = 0
        if addr:
            try:
                ip_num = int(addr.split('.')[-1].split('/')[0])
            except:
                pass
        return (disabled, iface, ip_num)
    
    sorted_peers = sorted(peers, key=sort_key)
    
    # Data rows
    for row, peer in enumerate(sorted_peers, start=2):
        # Row number
        ws_clients.cell(row=row, column=1).value = row - 1
        ws_clients.cell(row=row, column=1).style = 'data'
        ws_clients.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        
        # Display name
        ws_clients.cell(row=row, column=2).value = peer.get('DisplayName', '')
        ws_clients.cell(row=row, column=2).style = 'data'
        
        # User
        user = peer.get('UserIdentifier', '')
        ws_clients.cell(row=row, column=3).value = user
        ws_clients.cell(row=row, column=3).style = 'data'
        
        # IP Address
        addresses = peer.get('Addresses', [])
        if isinstance(addresses, list) and addresses:
            addr = addresses[0]
        else:
            addr = str(addresses)
        ws_clients.cell(row=row, column=4).value = addr
        ws_clients.cell(row=row, column=4).style = 'data'
        ws_clients.cell(row=row, column=4).font = Font(name='Consolas', size=10)
        
        # Interface
        iface = peer.get('Interface', '')
        cell = ws_clients.cell(row=row, column=5)
        cell.value = iface
        if 'wg0' in iface:
            cell.style = 'ma_vpn'
        else:
            cell.style = 'it_vpn'
        
        # Status
        disabled = peer.get('Disabled', False)
        cell = ws_clients.cell(row=row, column=6)
        cell.value = "DISABLED" if disabled else "Active"
        cell.style = 'disabled' if disabled else 'enabled'
        
        # Notes
        notes = peer.get('DisabledReason', '') or peer.get('Notes', '')
        ws_clients.cell(row=row, column=7).value = notes
        ws_clients.cell(row=row, column=7).style = 'data'
    
    # Add auto-filter
    ws_clients.auto_filter.ref = f"A1:G{len(sorted_peers) + 1}"
    
    # Freeze top row
    ws_clients.freeze_panes = 'A2'
    
    # ========== USERS SHEET ==========
    ws_users = wb.create_sheet("Users")
    
    # Headers
    user_headers = ['#', 'Email', 'First Name', 'Last Name', 'Admin', 'Status', 'Peer Count', 'Notes']
    for col, header in enumerate(user_headers, start=1):
        cell = ws_users.cell(row=1, column=col)
        cell.value = header
        cell.style = 'header'
    
    # Set column widths
    ws_users.column_dimensions['A'].width = 5
    ws_users.column_dimensions['B'].width = 35
    ws_users.column_dimensions['C'].width = 15
    ws_users.column_dimensions['D'].width = 15
    ws_users.column_dimensions['E'].width = 10
    ws_users.column_dimensions['F'].width = 12
    ws_users.column_dimensions['G'].width = 12
    ws_users.column_dimensions['H'].width = 35
    
    # Sort users: admins first, then active, then by email
    def user_sort_key(u):
        is_admin = 0 if u.get('IsAdmin') else 1
        disabled = 1 if u.get('Disabled') else 0
        email = u.get('Email', '').lower()
        return (is_admin, disabled, email)
    
    sorted_users_list = sorted(users, key=user_sort_key)
    
    # Data rows
    for row, user in enumerate(sorted_users_list, start=2):
        # Row number
        ws_users.cell(row=row, column=1).value = row - 1
        ws_users.cell(row=row, column=1).style = 'data'
        ws_users.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        
        # Email
        ws_users.cell(row=row, column=2).value = user.get('Email', '')
        ws_users.cell(row=row, column=2).style = 'data'
        
        # First name
        ws_users.cell(row=row, column=3).value = user.get('Firstname', '')
        ws_users.cell(row=row, column=3).style = 'data'
        
        # Last name
        ws_users.cell(row=row, column=4).value = user.get('Lastname', '')
        ws_users.cell(row=row, column=4).style = 'data'
        
        # Admin
        cell = ws_users.cell(row=row, column=5)
        cell.value = "Yes" if user.get('IsAdmin') else "No"
        if user.get('IsAdmin'):
            cell.font = Font(bold=True, color=COLORS['accent'])
        cell.style = 'data'
        cell.alignment = Alignment(horizontal='center')
        
        # Status
        disabled = user.get('Disabled', False)
        locked = user.get('Locked', False)
        status = "Active"
        if disabled:
            status = "Disabled"
        elif locked:
            status = "Locked"
        
        cell = ws_users.cell(row=row, column=6)
        cell.value = status
        if disabled:
            cell.style = 'disabled'
        elif locked:
            cell.font = Font(color=COLORS['warning'], bold=True)
            cell.fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
        else:
            cell.style = 'enabled'
        cell.alignment = Alignment(horizontal='center')
        
        # Peer count
        ws_users.cell(row=row, column=7).value = user.get('PeerCount', 0)
        ws_users.cell(row=row, column=7).style = 'data'
        ws_users.cell(row=row, column=7).alignment = Alignment(horizontal='center')
        
        # Notes
        notes = user.get('DisabledReason') or user.get('LockedReason') or user.get('Notes', '')
        ws_users.cell(row=row, column=8).value = notes
        ws_users.cell(row=row, column=8).style = 'data'
    
    # Add auto-filter
    ws_users.auto_filter.ref = f"A1:H{len(sorted_users_list) + 1}"
    
    # Freeze top row
    ws_users.freeze_panes = 'A2'
    
    # ========== SUMMARY SHEET ==========
    ws_summary = wb.create_sheet("Summary by User")
    
    # Create summary data
    user_summary = {}
    for peer in peers:
        user = peer.get('UserIdentifier', 'Unassigned') or 'Unassigned'
        if user not in user_summary:
            user_summary[user] = {'ma_vpn': 0, 'it_vpn': 0, 'active': 0, 'disabled': 0}
        
        if 'wg0' in peer.get('Interface', ''):
            user_summary[user]['ma_vpn'] += 1
        else:
            user_summary[user]['it_vpn'] += 1
        
        if peer.get('Disabled'):
            user_summary[user]['disabled'] += 1
        else:
            user_summary[user]['active'] += 1
    
    # Headers
    summary_headers = ['User', 'MA VPN Peers', 'IT VPN Peers', 'Total Active', 'Disabled', 'Total']
    for col, header in enumerate(summary_headers, start=1):
        cell = ws_summary.cell(row=1, column=col)
        cell.value = header
        cell.style = 'header'
    
    # Set column widths
    ws_summary.column_dimensions['A'].width = 35
    for col in 'BCDEF':
        ws_summary.column_dimensions[col].width = 15
    
    # Data
    sorted_summary = sorted(user_summary.items(), key=lambda x: -(x[1]['active'] + x[1]['disabled']))
    
    for row, (user, data) in enumerate(sorted_summary, start=2):
        ws_summary.cell(row=row, column=1).value = user
        ws_summary.cell(row=row, column=1).style = 'data'
        
        ws_summary.cell(row=row, column=2).value = data['ma_vpn']
        ws_summary.cell(row=row, column=2).style = 'data'
        ws_summary.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        
        ws_summary.cell(row=row, column=3).value = data['it_vpn']
        ws_summary.cell(row=row, column=3).style = 'data'
        ws_summary.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        
        ws_summary.cell(row=row, column=4).value = data['active']
        ws_summary.cell(row=row, column=4).style = 'enabled'
        
        ws_summary.cell(row=row, column=5).value = data['disabled']
        cell = ws_summary.cell(row=row, column=5)
        if data['disabled'] > 0:
            cell.style = 'disabled'
        else:
            cell.style = 'data'
            cell.alignment = Alignment(horizontal='center')
        
        ws_summary.cell(row=row, column=6).value = data['active'] + data['disabled']
        ws_summary.cell(row=row, column=6).style = 'data'
        ws_summary.cell(row=row, column=6).alignment = Alignment(horizontal='center')
        ws_summary.cell(row=row, column=6).font = Font(bold=True)
    
    # Add totals row
    total_row = len(sorted_summary) + 2
    ws_summary.cell(row=total_row, column=1).value = "TOTAL"
    ws_summary.cell(row=total_row, column=1).font = Font(bold=True, size=11)
    
    for col in range(2, 7):
        cell = ws_summary.cell(row=total_row, column=col)
        cell.value = f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{total_row-1})"
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color=COLORS['light'], end_color=COLORS['light'], fill_type='solid')
    
    # Freeze top row
    ws_summary.freeze_panes = 'A2'
    
    # Save workbook
    wb.save(output_file)
    print(f"Report saved to: {output_file}")
    
    # Print summary
    print(f"\n{'='*50}")
    print(f"VPN REPORT SUMMARY")
    print(f"{'='*50}")
    print(f"Total Peers:     {total_peers}")
    print(f"  - Active:      {active_peers}")
    print(f"  - Disabled:    {disabled_peers}")
    print(f"  - MA VPN:      {ma_vpn_peers}")
    print(f"  - IT VPN:      {it_vpn_peers}")
    print(f"Total Users:     {total_users}")
    print(f"  - Active:      {active_users}")
    print(f"{'='*50}")


def main():
    parser = argparse.ArgumentParser(description='Generate WireGuard VPN Report')
    parser.add_argument('-o', '--output', default='vpn-report.xlsx',
                        help='Output file path (default: vpn-report.xlsx)')
    parser.add_argument('--json', action='store_true',
                        help='Also export data as JSON')
    args = parser.parse_args()
    
    print("Loading secrets...")
    secrets = load_secrets()
    
    print("Fetching VPN data from wg-portal...")
    peers, users = fetch_vpn_data(secrets)
    
    print(f"Found {len(peers)} peers and {len(users)} users")
    
    print("Creating report...")
    create_report(peers, users, args.output)
    
    if args.json:
        json_file = args.output.rsplit('.', 1)[0] + '.json'
        with open(json_file, 'w') as f:
            json.dump({'peers': peers, 'users': users}, f, indent=2)
        print(f"JSON data saved to: {json_file}")


if __name__ == '__main__':
    main()
